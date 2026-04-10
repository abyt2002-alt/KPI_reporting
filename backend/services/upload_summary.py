from __future__ import annotations

import json
import os
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from pydantic import BaseModel, Field

try:
    from google import genai
    from google.genai import types
except Exception:  # pragma: no cover - graceful fallback when gemini sdk is unavailable
    genai = None
    types = None


LARGE_FILE_BYTES = int(os.getenv("UPLOAD_PROFILE_LARGE_FILE_BYTES", str(10 * 1024 * 1024)))
SAMPLE_ROWS = int(os.getenv("UPLOAD_PROFILE_SAMPLE_ROWS", "5000"))
PREVIEW_ROWS = int(os.getenv("UPLOAD_PROFILE_PREVIEW_ROWS", "50"))

FLASH_MODEL = os.getenv("GEMINI_FLASH_MODEL", "gemini-2.5-flash")
PRO_MODEL = os.getenv("GEMINI_PRO_MODEL", "gemini-2.5-pro")

DATE_HINTS = ("date", "time", "day", "month", "year", "timestamp", "datetime", "created", "updated")
ID_HINTS = ("id", "uuid", "guid", "key", "code", "identifier")
TEXT_HINTS = ("text", "description", "comment", "note", "message", "content", "body")
BOOL_VALUES = {"true", "false", "0", "1", "yes", "no", "y", "n", "t", "f"}
THEME_KEYWORDS = {
    "sales and marketing performance": (
        "sales",
        "revenue",
        "margin",
        "profit",
        "impression",
        "impressions",
        "click",
        "clicks",
        "campaign",
        "channel",
        "spend",
        "budget",
        "conversion",
        "orders",
        "order",
        "lead",
        "leads",
        "ctr",
        "cpc",
        "cpa",
        "roas",
    ),
    "ecommerce / product analytics": (
        "product",
        "products",
        "sku",
        "item",
        "items",
        "category",
        "brand",
        "store",
        "shop",
        "basket",
        "cart",
        "checkout",
        "customer",
        "customers",
        "session",
        "sessions",
        "traffic",
    ),
    "finance and profitability": (
        "expense",
        "expenses",
        "cost",
        "costs",
        "income",
        "revenue",
        "profit",
        "margin",
        "cash",
        "balance",
        "debt",
        "budget",
        "roi",
        "ebitda",
        "pnl",
        "loss",
    ),
    "operations and logistics": (
        "inventory",
        "shipment",
        "shipments",
        "delivery",
        "deliveries",
        "warehouse",
        "stock",
        "supply",
        "supplier",
        "lead time",
        "fulfillment",
        "logistics",
        "route",
    ),
    "customer analytics": (
        "customer",
        "customers",
        "user",
        "users",
        "segment",
        "segments",
        "cohort",
        "churn",
        "retention",
        "lifetime",
        "subscription",
        "loyalty",
        "engagement",
    ),
    "people and workforce": (
        "employee",
        "employees",
        "staff",
        "headcount",
        "hire",
        "hiring",
        "salary",
        "salarys",
        "department",
        "team",
        "role",
        "tenure",
    ),
}
THEME_ANGLE_HINTS = {
    "sales and marketing performance": [
        "Channel efficiency",
        "Spend vs revenue",
        "Conversion by channel",
        "Sales trend over time",
    ],
    "ecommerce / product analytics": [
        "Product mix",
        "Category performance",
        "Conversion by product",
        "Basket or order patterns",
    ],
    "finance and profitability": [
        "Profitability drivers",
        "Cost structure",
        "Margin trend",
        "Revenue vs expense",
    ],
    "operations and logistics": [
        "Throughput trend",
        "Delay hotspots",
        "Capacity utilization",
        "Process bottlenecks",
    ],
    "customer analytics": [
        "Segment behavior",
        "Retention patterns",
        "Churn drivers",
        "Cohort comparisons",
    ],
    "people and workforce": [
        "Headcount trend",
        "Department mix",
        "Compensation patterns",
        "Hiring or turnover drivers",
    ],
    "general business analytics": [
        "Correlation between drivers",
        "Trend over time",
        "Group comparison",
        "Outlier screening",
    ],
}


class SummaryOutput(BaseModel):
    title: str = Field(..., min_length=3, max_length=120)
    dataset_theme: str = Field(..., min_length=3, max_length=80)
    story: List[str] = Field(..., min_length=4, max_length=4)
    warnings: List[str] = Field(..., min_length=2, max_length=2)
    analysis_angles: List[str] = Field(..., min_length=4, max_length=4)


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _to_display_value(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (pd.Timedelta,)):
        return str(value)
    if pd.isna(value) if isinstance(value, float) else False:
        return None
    return value


def _clean_column_name(name: Any, index: int) -> str:
    if name is None or (isinstance(name, str) and not name.strip()):
        return f"column_{index + 1}"
    return str(name).strip()


def _detect_file_type(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv"}:
        return "csv"
    if suffix in {".xlsx", ".xls"}:
        return "excel"
    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


def _estimate_csv_rows(content: bytes) -> int:
    if not content:
        return 0
    return max(content.count(b"\n"), 1)


def _read_csv_sample(content: bytes, sample_rows: int) -> pd.DataFrame:
    return pd.read_csv(
        BytesIO(content),
        low_memory=False,
        nrows=sample_rows,
    )


def _read_excel_sample(content: bytes, sample_rows: int) -> Tuple[pd.DataFrame, int, str]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError("The first sheet is empty.")

    headers = [_clean_column_name(col, idx) for idx, col in enumerate(header)]
    sample_data: List[Dict[str, Any]] = []
    for row in rows:
        if len(sample_data) >= sample_rows:
            break
        record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        sample_data.append(record)

    df = pd.DataFrame(sample_data, columns=headers)
    total_rows = max(int(sheet.max_row or 1) - 1, len(df))
    workbook.close()
    return df, total_rows, sheet.title or "Sheet1"


def _is_bool_like(values: pd.Series) -> bool:
    if values.empty:
        return False
    normalized = values.astype(str).str.strip().str.lower()
    return normalized.isin(BOOL_VALUES).mean() >= 0.9


def _has_datetime_name_hint(name: str) -> bool:
    lower_name = name.lower()
    tokens = [token for token in re.split(r"[^a-z0-9]+", lower_name) if token]
    if any(token in {"date", "time", "timestamp", "datetime", "created", "updated", "year", "month"} for token in tokens):
        return True
    return lower_name.endswith(("date", "time", "timestamp", "datetime", "created_at", "updated_at"))


def _looks_datetime_like(name: str, values: pd.Series) -> bool:
    if values.empty:
        return False

    lower_name = name.lower()
    name_hint = _has_datetime_name_hint(lower_name)
    raw_strings = values.astype(str).str.strip()
    digit_only_ratio = raw_strings.str.fullmatch(r"\d+").mean()
    if digit_only_ratio >= 0.9 and not name_hint:
        return False

    sample_strings = raw_strings.head(25)
    if sample_strings.empty:
        return False

    parsed = _parse_datetime_series(values)
    parse_ratio = parsed.notna().mean()
    if parse_ratio < 0.8:
        return False

    parsed_non_null = parsed.dropna()
    if parsed_non_null.nunique(dropna=True) < 2:
        return False

    if values.dtype.kind in {"i", "u", "f"} and not name_hint:
        return False

    separator_hits = sample_strings.str.contains(r"[-/:T ]", regex=True).mean()
    return name_hint or separator_hits >= 0.35 or parse_ratio >= 0.95


def _parse_datetime_series(values: pd.Series) -> pd.Series:
    return pd.to_datetime(values, errors="coerce", format="mixed")


def _infer_semantic_type(name: str, raw_dtype: str, values: pd.Series) -> Tuple[str, str]:
    non_null = values.dropna()
    if non_null.empty:
        return "unknown", "Column has no non-null values in the sample."

    normalized_name = name.lower()
    unique_ratio = float(non_null.nunique(dropna=True)) / max(len(non_null), 1)
    avg_len = float(non_null.astype(str).str.len().mean())

    if _looks_datetime_like(name, non_null):
        dt_ratio = _parse_datetime_series(non_null).notna().mean()
        if dt_ratio >= 0.8:
            return "datetime", "Date-like values with a clear time axis signal."

    if _is_bool_like(non_null):
        return "boolean", "Values look boolean-like."

    numeric = pd.to_numeric(non_null, errors="coerce")
    numeric_ratio = numeric.notna().mean() if len(non_null) else 0.0

    if numeric_ratio >= 0.85:
        if any(hint in normalized_name for hint in ID_HINTS):
            return "id", "Numeric column with ID-like uniqueness."
        return "numeric", "Numeric values dominate the sample."

    if any(hint in normalized_name for hint in ID_HINTS) and unique_ratio >= 0.7:
        return "id", "Column name and uniqueness suggest an identifier."

    if avg_len >= 40:
        return "text", "Free-text-like values with long strings."

    if unique_ratio >= 0.9 and len(non_null) >= 20:
        return "id", "High-cardinality categorical values that behave like IDs."

    if any(hint in normalized_name for hint in TEXT_HINTS):
        return "text", "Text-heavy column name and sample values."

    return "categorical", "Mixed or low-cardinality values."


def _compact_numeric_summary(series: pd.Series) -> Optional[Dict[str, float]]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return None
    return {
        "mean": round(float(numeric.mean()), 6),
        "std": round(float(numeric.std(ddof=0)), 6) if len(numeric) > 1 else 0.0,
        "min": round(float(numeric.min()), 6),
        "p25": round(float(numeric.quantile(0.25)), 6),
        "median": round(float(numeric.quantile(0.5)), 6),
        "p75": round(float(numeric.quantile(0.75)), 6),
        "max": round(float(numeric.max()), 6),
    }


def _compact_datetime_summary(series: pd.Series) -> Optional[Dict[str, str]]:
    parsed = _parse_datetime_series(series.dropna())
    parsed = parsed.dropna()
    if parsed.empty:
        return None
    dt_min = parsed.min()
    dt_max = parsed.max()
    return {"min": dt_min.isoformat(), "max": dt_max.isoformat()}


def _compact_top_values(series: pd.Series, limit: int = 3) -> List[Dict[str, Any]]:
    non_null = series.dropna().astype(str).str.strip()
    if non_null.empty:
        return []
    counts = non_null.value_counts().head(limit)
    return [{"value": str(idx), "count": int(count)} for idx, count in counts.items()]


def _tokenize_for_theme(text: str) -> List[str]:
    return [token for token in re.split(r"[^a-z0-9]+", text.lower()) if token]


def _score_theme_name(name: str, keywords: Tuple[str, ...]) -> int:
    lower_name = name.lower()
    score = 0
    for keyword in keywords:
        if " " in keyword:
            if keyword in lower_name:
                score += 3
        else:
            token_hits = _tokenize_for_theme(lower_name).count(keyword)
            if token_hits:
                score += 2 * token_hits
            elif keyword in lower_name:
                score += 1
    return score


def _extract_theme_candidates(columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    for theme, keywords in THEME_KEYWORDS.items():
        score = 0
        matched_columns: List[str] = []
        matched_terms: List[str] = []
        for column in columns:
            column_score = _score_theme_name(column["n"], keywords)
            if column_score > 0:
                score += column_score
                matched_columns.append(column["n"])
        if score > 0:
            for keyword in keywords:
                if any(keyword in column["n"].lower() for column in columns):
                    matched_terms.append(keyword)
            candidates.append(
                {
                    "theme": theme,
                    "score": score,
                    "columns": matched_columns[:5],
                    "terms": matched_terms[:6],
                }
            )

    candidates.sort(key=lambda item: item["score"], reverse=True)
    return candidates


def _infer_dataset_theme(columns: List[Dict[str, Any]], groups: Dict[str, List[str]]) -> Dict[str, Any]:
    candidates = _extract_theme_candidates(columns)
    if not candidates:
        return {
            "label": "general business analytics",
            "confidence": 0,
            "reasons": ["No strong business-specific keywords were detected."],
            "seed_angles": THEME_ANGLE_HINTS["general business analytics"][:3],
            "signal_columns": [],
        }

    top = candidates[0]
    secondary = candidates[1] if len(candidates) > 1 else None

    label = top["theme"]
    if secondary and secondary["score"] >= max(2, int(top["score"] * 0.7)):
        if {top["theme"], secondary["theme"]} == {"sales and marketing performance", "ecommerce / product analytics"}:
            label = "ecommerce sales and marketing"
        elif {top["theme"], secondary["theme"]} == {"sales and marketing performance", "finance and profitability"}:
            label = "commercial performance"
        else:
            label = f"{top['theme']} + {secondary['theme']}"

    reasons = [
        f"Matched columns: {', '.join(top['columns'][:3])}" if top["columns"] else "Matched business keywords in column names.",
    ]
    if top["terms"]:
        reasons.append(f"Signal terms: {', '.join(top['terms'][:4])}.")
    if secondary:
        reasons.append(f"Secondary signal: {secondary['theme']}.")
    if groups.get("datetime"):
        reasons.append("A time axis is present, so trend analysis is relevant.")

    seed_angles = list(THEME_ANGLE_HINTS.get(top["theme"], THEME_ANGLE_HINTS["general business analytics"]))
    if secondary:
        for angle in THEME_ANGLE_HINTS.get(secondary["theme"], []):
            if angle not in seed_angles:
                seed_angles.append(angle)
            if len(seed_angles) >= 5:
                break
    if len(seed_angles) < 3:
        for angle in THEME_ANGLE_HINTS["general business analytics"]:
            if angle not in seed_angles:
                seed_angles.append(angle)
            if len(seed_angles) >= 3:
                break

    signal_columns: List[str] = []
    for candidate in candidates[:2]:
        for column_name in candidate["columns"]:
            if column_name not in signal_columns:
                signal_columns.append(column_name)

    return {
        "label": label,
        "confidence": top["score"],
        "reasons": reasons[:4],
        "seed_angles": seed_angles[:4],
        "signal_columns": signal_columns[:6],
    }


def _score_narrative_column(column: Dict[str, Any], time_column: Optional[str]) -> float:
    score = 0.0
    if column["n"] == time_column:
        score += 6.0
    if column["t"] == "datetime":
        score += 5.0
    elif column["t"] == "numeric":
        score += 4.0
    elif column["t"] == "categorical":
        score += 3.0
    elif column["t"] == "boolean":
        score += 2.0
    elif column["t"] in {"id", "text"}:
        score -= 2.5
    score += max(0.0, 1.0 - float(column.get("null", 0.0)) * 1.5)
    score += max(0.0, 1.0 - abs(float(column.get("uniq", 0.0)) - 0.5))
    return score


def _build_narrative_context(
    columns: List[Dict[str, Any]],
    groups: Dict[str, List[str]],
    theme: Dict[str, Any],
    row_count: int,
    sample_row_count: int,
) -> Dict[str, Any]:
    time_column = _choose_time_column(columns)
    ranked_columns = sorted(columns, key=lambda c: _score_narrative_column(c, time_column), reverse=True)

    focus_columns: List[Dict[str, Any]] = []
    for column in ranked_columns:
        if column["t"] in {"id"}:
            continue
        if column["n"] in {item["name"] for item in focus_columns}:
            continue
        focus_columns.append(
            {
                "name": column["n"],
                "type": column["t"],
                "note": column.get("note"),
                "samples": column.get("s", [])[:3],
            }
        )
        if len(focus_columns) >= 6:
            break

    key_groups = {
        "numeric": groups.get("numeric", [])[:5],
        "categorical": groups.get("categorical", [])[:5],
        "datetime": groups.get("datetime", [])[:3],
    }

    return {
        "dataset_focus": theme.get("label", "general business analytics"),
        "confidence": theme.get("confidence", 0),
        "signals": theme.get("reasons", []),
        "signal_columns": theme.get("signal_columns", []),
        "key_groups": key_groups,
        "focus_columns": focus_columns,
        "row_count": row_count,
        "sample_row_count": sample_row_count,
    }


def _build_column_profile(name: str, series: pd.Series) -> Dict[str, Any]:
    non_null = series.dropna()
    total = len(series)
    non_null_count = int(non_null.shape[0])
    missing_count = int(total - non_null_count)
    unique_count = int(non_null.nunique(dropna=True)) if non_null_count else 0
    unique_pct = round(unique_count / non_null_count, 4) if non_null_count else 0.0
    missing_pct = round(missing_count / total, 4) if total else 0.0
    raw_dtype = str(series.dtype)
    semantic_type, reason = _infer_semantic_type(name, raw_dtype, series)
    sample_values = [_to_display_value(v) for v in non_null.head(3).tolist()]

    profile: Dict[str, Any] = {
        "n": name,
        "d": raw_dtype,
        "t": semantic_type,
        "null": missing_pct,
        "uniq": unique_pct,
        "note": reason,
        "s": sample_values,
    }

    high_cardinality = unique_pct >= 0.5 or unique_count >= 50

    if semantic_type == "numeric":
        if not high_cardinality:
            profile["stats"] = _compact_numeric_summary(series)
        else:
            profile["note"] = "High-cardinality numeric column; distribution stats trimmed."
    elif semantic_type == "datetime":
        profile["range"] = _compact_datetime_summary(series)
    elif semantic_type == "categorical" and not high_cardinality:
        profile["top"] = _compact_top_values(series, limit=3)
    elif semantic_type in {"id", "text"}:
        profile["note"] = "High-cardinality or text-heavy column; trimmed to compact metadata."

    return profile


def _choose_time_column(columns: List[Dict[str, Any]]) -> Optional[str]:
    candidates = [c for c in columns if c["t"] == "datetime"]
    if not candidates:
        return None
    candidates.sort(key=lambda c: (float(c.get("null", 0.0)), -float(c.get("uniq", 0.0))))
    return candidates[0]["n"]


def _build_fallback_summary(profile: Dict[str, Any]) -> SummaryOutput:
    shape = profile["shape"]
    groups = profile["groups"]
    theme = profile.get("theme", {})
    theme_label = theme.get("label", "general business analytics")
    signal_columns = theme.get("signal_columns", [])
    time_info = profile.get("time", {})
    time_column = time_info.get("column")
    date_range = time_info.get("range")
    numeric_count = len(groups.get("numeric", []))
    categorical_count = len(groups.get("categorical", []))
    datetime_count = len(groups.get("datetime", []))

    story = [
        f"This dataset has {shape['rows']:,} rows and {shape['cols']} columns.",
        f"It looks most like {theme_label}.",
        f"It contains {numeric_count} numeric, {categorical_count} categorical, and {datetime_count} datetime columns.",
        (
            f"{time_column} spans {date_range['min']} to {date_range['max']}."
            if time_column and date_range
            else f"{time_column} looks like the main time axis."
            if time_column
            else (
                f"Key fields include {', '.join(signal_columns[:3])}."
                if signal_columns
                else "No strong time column signal was detected."
            )
        ),
    ]

    warnings = []
    if groups.get("id"):
        warnings.append(f"Possible ID columns: {', '.join(groups['id'][:3])}.")
    if groups.get("text"):
        warnings.append(f"Text-heavy columns may need separate preprocessing: {', '.join(groups['text'][:3])}.")
    if profile.get("quality", {}).get("duplicate_rows"):
        warnings.append(f"Duplicate rows were observed in the sample: {profile['quality']['duplicate_rows']}.")
    if not warnings:
        warnings.append("No major structural issue was detected in the sample.")
    if len(warnings) < 2:
        warnings.append("Check sparse columns and mixed dtypes before modeling.")

    analysis_angles = list(theme.get("seed_angles", []))
    if time_column and "Trend over time" not in analysis_angles and "Daily sales trend" not in analysis_angles:
        analysis_angles.insert(0, "Trend over time")
    if numeric_count >= 2 and not any("correlation" in angle.lower() or "regression" in angle.lower() for angle in analysis_angles):
        analysis_angles.append("Correlation between drivers")
    if categorical_count >= 1 and not any("group" in angle.lower() or "segment" in angle.lower() for angle in analysis_angles):
        analysis_angles.append("Group comparison")
    if len(analysis_angles) < 4:
        analysis_angles.append("Basic column profiling")
    if len(analysis_angles) < 4:
        analysis_angles.append("Outlier screening")

    return SummaryOutput(
        title=f"{theme_label.title()} snapshot",
        dataset_theme=theme_label.title(),
        story=story[:4],
        warnings=warnings[:2],
        analysis_angles=analysis_angles[:4],
    )


def _build_prompt_profile(
    filename: str,
    file_type: str,
    profiling_mode: str,
    row_count: int,
    sample_row_count: int,
    columns: List[Dict[str, Any]],
    groups: Dict[str, List[str]],
    time_column: Optional[str],
    date_range: Optional[Dict[str, Any]],
    duplicate_rows: Optional[int],
    theme: Dict[str, Any],
    narrative: Dict[str, Any],
) -> Dict[str, Any]:
    return {
        "file": {"name": filename, "type": file_type, "mode": profiling_mode},
        "shape": {
            "rows": row_count,
            "cols": len(columns),
            "sample_rows": sample_row_count,
            "estimated": profiling_mode == "sampled",
        },
        "groups": groups,
        "time": {"column": time_column, "range": date_range},
        "quality": {"duplicate_rows": duplicate_rows},
        "theme": theme,
        "narrative": narrative,
        "columns": columns,
    }


def _call_gemini_summary(profile: Dict[str, Any], column_count: int) -> Tuple[SummaryOutput, str]:
    if genai is None or types is None:
        raise RuntimeError("Gemini SDK is not available")

    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY is not set")

    model = FLASH_MODEL if column_count <= 30 else PRO_MODEL
    client = genai.Client(api_key=api_key)

    system_instruction = (
        "You are a concise data analyst for a spreadsheet app.\n"
        "Use only the provided JSON profile.\n"
        "Return valid JSON only.\n"
        "First identify the most likely dataset theme from the profile and column names.\n"
        "Prefer a business lens such as sales, marketing, ecommerce, finance, operations, customer, or workforce when the signals support it.\n"
        "Use the narrative context and actual field names to explain what the file looks like, as if you were naming the dataset for a dashboard.\n"
        "If there are strong sales or marketing signals, discuss channel efficiency, spend allocation, conversion, revenue, impressions, and trend over time.\n"
        "If there are product or ecommerce signals, discuss product mix, category performance, and sales contribution.\n"
        "If the theme is weak, use general business analytics.\n"
        "Write a title of 3-6 words.\n"
        "Write a short dataset theme label, 2-4 words, in title case.\n"
        "Write exactly 4 story bullets, each 16 words or fewer, and each one must be unique.\n"
        "Write exactly 2 warnings, each 12 words or fewer, and each one must be unique.\n"
        "Write exactly 4 analysis angles, each 4 words or fewer, and each one must be unique.\n"
        "Do not invent facts, do not repeat yourself, and do not add markdown or extra prose."
    )
    user_prompt = f"PROFILE_JSON={json.dumps(profile, separators=(',', ':'), ensure_ascii=False, default=str)}"

    config = types.GenerateContentConfig(
        systemInstruction=system_instruction,
        temperature=0.2,
        responseMimeType="application/json",
        responseSchema=SummaryOutput,
    )

    response = client.models.generate_content(
        model=model,
        contents=[user_prompt],
        config=config,
    )

    raw_text = response.text or ""
    parsed: Dict[str, Any] = json.loads(raw_text)
    summary = SummaryOutput.model_validate(parsed)
    return summary, model


def _compact_preview_rows(df: pd.DataFrame, limit: int = PREVIEW_ROWS) -> List[Dict[str, Any]]:
    preview = df.head(limit).replace({pd.NaT: None})
    records: List[Dict[str, Any]] = []
    for row in preview.to_dict(orient="records"):
        cleaned: Dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, (pd.Timestamp, datetime)):
                cleaned[str(key)] = value.isoformat()
            elif value is None or (isinstance(value, float) and pd.isna(value)):
                cleaned[str(key)] = None
            else:
                cleaned[str(key)] = value
        records.append(cleaned)
    return records


def _profile_dataframe(
    df: pd.DataFrame,
    filename: str,
    file_type: str,
    profiling_mode: str,
    row_count: int,
    sample_row_count: int,
) -> Dict[str, Any]:
    df = df.copy()
    df.columns = [_clean_column_name(col, idx) for idx, col in enumerate(df.columns)]

    columns: List[Dict[str, Any]] = []
    groups: Dict[str, List[str]] = {
        "numeric": [],
        "categorical": [],
        "datetime": [],
        "boolean": [],
        "text": [],
        "id": [],
        "unknown": [],
    }

    for column in df.columns:
        profile = _build_column_profile(column, df[column])
        columns.append(profile)
        groups.setdefault(profile["t"], []).append(profile["n"])

    theme = _infer_dataset_theme(columns, groups)
    narrative = _build_narrative_context(columns, groups, theme, row_count, sample_row_count)
    time_column = _choose_time_column(columns)
    date_range: Optional[Dict[str, Any]] = None
    if time_column:
        dt_profile = next((c for c in columns if c["n"] == time_column), None)
        if dt_profile and dt_profile.get("range"):
            date_range = dt_profile["range"]

    duplicate_rows: Optional[int]
    if profiling_mode == "full":
        duplicate_rows = int(df.duplicated().sum())
    else:
        sample_duplicates = int(df.duplicated().sum())
        duplicate_rows = sample_duplicates if sample_duplicates > 0 else None

    prompt_profile = _build_prompt_profile(
        filename=filename,
        file_type=file_type,
        profiling_mode=profiling_mode,
        row_count=row_count,
        sample_row_count=sample_row_count,
        columns=columns,
        groups=groups,
        time_column=time_column,
        date_range=date_range,
        duplicate_rows=duplicate_rows,
        theme=theme,
        narrative=narrative,
    )

    try:
        summary, model_used = _call_gemini_summary(prompt_profile, len(columns))
        used_fallback = False
    except Exception:
        summary = _build_fallback_summary(prompt_profile)
        model_used = FLASH_MODEL if len(columns) <= 30 else PRO_MODEL
        used_fallback = True

    numeric_columns = groups.get("numeric", [])
    categorical_columns = groups.get("categorical", [])
    datetime_columns = groups.get("datetime", [])

    return {
        "filename": filename,
        "file_type": file_type,
        "profiling_mode": profiling_mode,
        "row_count": row_count,
        "row_count_is_estimated": profiling_mode == "sampled",
        "column_count": len(columns),
        "sample_row_count": sample_row_count,
        "sample_rows": _compact_preview_rows(df),
        "columns": columns,
        "column_groups": {
            "numeric": numeric_columns,
            "categorical": categorical_columns,
            "datetime": datetime_columns,
            "boolean": groups.get("boolean", []),
            "text": groups.get("text", []),
            "id": groups.get("id", []),
            "unknown": groups.get("unknown", []),
        },
        "time_column": time_column,
        "date_range": date_range,
        "dataset_theme": summary.dataset_theme,
        "summary": summary.model_dump(),
        "model_used": model_used,
        "used_fallback": used_fallback,
    }


def analyze_uploaded_file(filename: str, content: bytes) -> Dict[str, Any]:
    file_type = _detect_file_type(filename)
    file_size = len(content)

    if file_type == "csv":
        row_count_estimate = max(_estimate_csv_rows(content) - 1, 0)
        if row_count_estimate <= SAMPLE_ROWS and file_size <= LARGE_FILE_BYTES:
            df = pd.read_csv(BytesIO(content), low_memory=False)
            return _profile_dataframe(df, filename, file_type, "full", len(df), len(df))

        df = _read_csv_sample(content, SAMPLE_ROWS)
        return _profile_dataframe(
            df,
            filename,
            file_type,
            "sampled",
            row_count_estimate,
            len(df),
        )

    df, row_count, _sheet_name = _read_excel_sample(content, SAMPLE_ROWS)
    profiling_mode = "full" if row_count <= SAMPLE_ROWS and file_size <= LARGE_FILE_BYTES else "sampled"
    return _profile_dataframe(df, filename, file_type, profiling_mode, row_count, len(df))
